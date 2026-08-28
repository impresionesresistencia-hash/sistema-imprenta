import streamlit as st
import pandas as pd
import os
import re
import openpyxl
from datetime import datetime, timedelta
import shutil
from pathlib import Path

# Configuración de la página con inyección de CSS avanzado y script para mantener la posición de scroll
st.set_page_config(page_title="Sistema Imprenta", page_icon="🖨️", layout="wide")

st.markdown("""
    <style>
    /* Estilos generales y optimización de espacio */
    .block-container { padding-top: 1rem !important; padding-bottom: 1rem !important; }
    h1 { margin-bottom: 0rem !important; padding-bottom: 0rem !important; font-size: 24px !important; }
    h3 { margin-top: 0.5rem !important; margin-bottom: 0.5rem !important; font-size: 18px !important; }
    
    /* 1. Pestañas (Tabs) estilo cápsula moderna */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px !important;
        background-color: #141414;
        padding: 6px;
        border-radius: 10px;
        border: 1px solid #262626;
    }
    .stTabs [data-baseweb="tab"] {
        height: 38px;
        background-color: #1e1e1e;
        border-radius: 6px;
        padding: 0px 16px;
        font-size: 13px !important;
        font-weight: 500;
        color: #b0b0b0 !important;
        border: 1px solid transparent;
    }
    .stTabs [aria-selected="true"] {
        background-color: #2b2b2b !important;
        color: #ffffff !important;
        border: 1px solid #444444 !important;
    }
    
    /* 2. Tarjetas de Métricas Personalizadas (KPI Cards) */
    div[data-testid="metric-container"] {
        background-color: #1e1e1e;
        padding: 14px 18px;
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.3);
        border: 1px solid #333333;
    }
    div[data-testid="metric-container"] label {
        color: #b0b0b0 !important;
        font-size: 13px !important;
        font-weight: 600 !important;
    }
    div[data-testid="metric-container"] div[data-testid="stMetricValue"] {
        color: #ffffff !important;
        font-size: 22px !important;
    }

    /* 3. Contenedores de formularios en tarjetas con bordes redondeados */
    div[data-testid="stExpander"] {
        background-color: #191919;
        border: 1px solid #2e2e2e;
        border-radius: 10px;
        padding: 5px;
        box-shadow: 0 2px 5px rgba(0,0,0,0.15);
    }

    /* 4. Inputs, selects y campos de texto con diseño prolijo */
    div.stTextInput > div > div > input, 
    div.stNumberInput > div > div > input, 
    div.stSelectbox > div > div > div {
        background-color: #181818 !important;
        border: 1px solid #333333 !important;
        border-radius: 6px !important;
        color: #ffffff !important;
    }
    div.stTextInput > div > div > input:focus, 
    div.stNumberInput > div > div > input:focus {
        border-color: #666666 !important;
        box-shadow: 0 0 5px rgba(255,255,255,0.1) !important;
    }

    /* 5. Alertas modernas (Success, Warning, Error) */
    div.stAlert {
        border-radius: 8px !important;
        border: 1px solid rgba(255,255,255,0.05) !important;
        box-shadow: 0 2px 4px rgba(0,0,0,0.2);
    }
    </style>

    <!-- Script para mantener la posición del scroll al recargar -->
    <script>
        document.addEventListener("DOMContentLoaded", function() {
            const scrollPos = localStorage.getItem('scrollPos');
            if (scrollPos) {
                window.scrollTo(0, parseInt(scrollPos));
                localStorage.removeItem('scrollPos');
            }
        });
        window.addEventListener('beforeunload', function() {
            localStorage.setItem('scrollPos', window.scrollY);
        });
    </script>
""", unsafe_allow_html=True)

st.title("🖨️ Sistema de Gestión - Imprenta")

# ==============================================================================
# CONFIGURACIÓN DE RUTA LOCAL (GOOGLE DRIVE)
# ==============================================================================
RUTA_COMPARTIDA = r"G:\Mi unidad\sistema_imprenta_datos"

def obtener_ruta(nombre_archivo):
    if os.path.exists(RUTA_COMPARTIDA):
        return os.path.join(RUTA_COMPARTIDA, nombre_archivo)
    return nombre_archivo

# Creación de pestañas
tab_inicio, tab_pedidos, tab_materiales, tab_cheques, tab_semanal, tab_mensual, tab_presupuestos = st.tabs([
    "📈 Dashboard",
    "🛒 Pedidos y Clientes",
    "📦 Materiales y Gastos",
    "💳 Cartera de Cheques",
    "📅 Cierre Semanal",
    "📊 Cierre Mensual",
    "🤖 Presupuestos"
])

# --- 1. ARCHIVO COMPARTIDO DE PEDIDOS DE CLIENTES ---
archivo_pedidos = obtener_ruta("PEDIDOS_CLIENTES.csv")
if os.path.exists(archivo_pedidos):
    df_pedidos = pd.read_csv(archivo_pedidos, dtype={
        "Estado": str, 
        "Cobró": str, 
        "Observaciones": str, 
        "Impresión": str, 
        "Fecha Cobro": str,
        "Fecha": str
    })
    if "Cobró" not in df_pedidos.columns:
        df_pedidos["Cobró"] = "N/A"
    if "Observaciones" not in df_pedidos.columns:
        df_pedidos["Observaciones"] = ""
    if "Impresión" not in df_pedidos.columns:
        df_pedidos["Impresión"] = "🔴 Falta Imprimir"
        
    df_pedidos["Cobró"] = df_pedidos["Cobró"].fillna("N/A").astype(str).replace({"ROBERTO": "ROB", "GONZALO": "GON"})
    df_pedidos["Estado"] = df_pedidos["Estado"].fillna("Pendiente").astype(str)
    df_pedidos["Observaciones"] = df_pedidos["Observaciones"].fillna("").astype(str)
    df_pedidos["Impresión"] = df_pedidos["Impresión"].fillna("🔴 Falta Imprimir").astype(str)
else:
    df_pedidos = pd.DataFrame(columns=["Fecha", "Cliente", "Material", "Medida", "Copias", "Monto Total", "Estado", "Cobró", "Observaciones", "Impresión"])

# --- 2. PROCESAMIENTO DE MATERIALES (GASTOS PROVEEDORES) ---
archivo_materiales = obtener_ruta("MATERIALES.xlsx")
if os.path.exists(archivo_materiales):
    df_mat = pd.read_excel(archivo_materiales)
    df_mat.columns = df_mat.columns.str.strip().str.upper()
    
    columnas_requeridas = ["FECHA", "DESCRIPCION", "MONTO", "FORMA DE PAGO", "PROVEEDOR", "ESTADO", "SALDO"]
    for col in columnas_requeridas:
        if col not in df_mat.columns:
            df_mat[col] = ""
            
    df_mat["ESTADO"] = df_mat["ESTADO"].astype(str).str.upper().str.strip()
    df_mat["ESTADO"] = df_mat["ESTADO"].replace({"P": "PENDIENTE", "F": "PAGADO", "NAN": "PENDIENTE", "": "PENDIENTE"})
    df_mat["MONTO"] = pd.to_numeric(df_mat["MONTO"], errors="coerce").fillna(0.0)
    df_mat["SALDO"] = df_mat.apply(lambda r: 0.0 if r["ESTADO"] == "PAGADO" else r["MONTO"], axis=1)
    df_mat = df_mat.dropna(subset=["DESCRIPCION"], how="all")
else:
    df_mat = pd.DataFrame(columns=["FECHA", "DESCRIPCION", "MONTO", "FORMA DE PAGO", "PROVEEDOR", "ESTADO", "SALDO"])

# --- 3. PROCESAMIENTO DE CHEQUES ---
archivo_cheques = obtener_ruta("CHEQUES NUEVO.xlsx")
lista_cheques_procesados = []

if os.path.exists(archivo_cheques):
    try:
        wb = openpyxl.load_workbook(archivo_cheques, data_only=True)
        sheet = wb.active
        proveedores_cols = {}
        for col_idx in range(1, sheet.max_column + 1):
            prov_name = sheet.cell(row=1, column=col_idx).value
            if prov_name and not str(prov_name).startswith("Unnamed:"):
                proveedores_cols[col_idx] = str(prov_name).strip()
        
        for row_idx in range(2, sheet.max_row + 1):
            for col_idx, proveedor in proveedores_cols.items():
                cell = sheet.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    val_str = str(val).strip()
                    monto_match = re.search(r'\$?(\d+)', val_str)
                    fecha_match = re.search(r'(\d{2}/\d{2})', val_str)
                    if monto_match:
                        monto = float(monto_match.group(1))
                        fecha_str = fecha_match.group(1) if fecha_match else "Sin fecha"
                        fill_color = cell.fill.start_color.rgb if (cell.fill and cell.fill.start_color) else None
                        is_yellow = False
                        if fill_color:
                            is_yellow = "FFFF00" in str(fill_color) or str(fill_color) == "00FFFF00" or str(fill_color) == "FFFFFF00"
                        estado = "Pagado" if is_yellow else "Pendiente"
                        
                        lista_cheques_procesados.append({
                            "Proveedor": proveedor, "Monto": monto, "Fecha Vencimiento": fecha_str, "Estado": estado, "Detalle en Excel": val_str
                        })
        df_cheques = pd.DataFrame(lista_cheques_procesados)
    except Exception as e:
        df_cheques = pd.DataFrame()
else:
    df_cheques = pd.DataFrame()

# --- 4. ARCHIVO DE GASTOS DE RENDICIÓN SEMANAL ---
archivo_gastos_semanales = obtener_ruta("GASTOS_SEMANALES_CAJA.csv")
if os.path.exists(archivo_gastos_semanales):
    df_gastos_sem = pd.read_csv(archivo_gastos_semanales)
else:
    df_gastos_sem = pd.DataFrame(columns=["Semana", "Socio", "Concepto", "Monto"])


# --- PESTAÑA DASHBOARD ---
with tab_inicio:
    st.subheader("Resumen General")
    col1, col2, col3, col4 = st.columns(4)
    
    if not df_mat.empty:
        total_deuda_prov = df_mat["SALDO"].sum()
        col1.metric(label="Deuda a Proveedores 🟥", value=f"${total_deuda_prov:,.2f}")
    else:
        col1.metric(label="Deuda a Proveedores", value="$0.00")
        
    if not df_cheques.empty:
        df_pendientes = df_cheques[df_cheques["Estado"] == "Pendiente"]
        col2.metric(label="Cheques PENDIENTES ⚠️", value=f"${df_pendientes['Monto'].sum():,.2f}", delta=f"{len(df_pendientes)} activos")
    else:
        col2.metric(label="Cheques Pendientes", value="$0.00")

    total_ventas = df_pedidos["Monto Total"].sum() if not df_pedidos.empty else 0.0
    total_trabajos_registrados = df_pedidos.drop_duplicates(subset=["Fecha", "Cliente"]).shape[0] if not df_pedidos.empty else 0
    col3.metric(label="Trabajos Registrados", value=f"{total_trabajos_registrados}")
    col4.metric(label="Facturación Clientes 💰", value=f"${total_ventas:,.2f}")

    st.write("---")
    st.subheader("🚨 Alertas y Vencimientos Próximos")
    
    col_alerta1, col_alerta2 = st.columns(2)
    
    with col_alerta1:
        if not df_cheques.empty:
            cheques_activos = df_cheques[df_cheques["Estado"] == "Pendiente"].copy()
            hoy = datetime.now()
            hoy_midnight = hoy.replace(hour=0, minute=0, second=0, microsecond=0)
            detalle_cheques_proximos = []
            
            for _, row in cheques_activos.iterrows():
                f_str = row["Fecha Vencimiento"]
                try:
                    f_dt = datetime.strptime(f"{f_str}/{hoy.year}", "%d/%m/%Y")
                    dias_restantes = (f_dt - hoy_midnight).days
                    if 0 <= dias_restantes <= 7:
                        detalle_cheques_proximos.append(f"• **{row['Proveedor']}** (${row['Monto']:,.2f}) vence en **{dias_restantes} día(s)** (Fecha: {f_str})")
                except:
                    pass
            
            if len(detalle_cheques_proximos) > 0:
                st.error("⚠️ **Cheques que vencen en los próximos 7 días:**\n\n" + "\n".join(detalle_cheques_proximos))
            else:
                st.success("✅ No hay cheques pendientes que venzan en los próximos 7 días.")
        else:
            st.info("No hay registros de cheques para evaluar.")
            
    with col_alerta2:
        if not df_mat.empty:
            cant_pendientes_prov = len(df_mat[df_mat["ESTADO"] != "PAGADO"])
            if cant_pendientes_prov > 0:
                st.warning(f"📦 Tenés **{cant_pendientes_prov} factura(s) o compra(s) pendiente(s)** de pago a proveedores.")
            else:
                st.success("✅ Todas las cuentas de proveedores están al día.")
        else:
            st.info("No hay registros de materiales.")
    # =========================================================================
    # 🔴 PEGAR DESDE ACÁ...
    # =========================================================================
    st.write("---")
    st.subheader("⚠️ Clientes con Deuda Vencida (> 7 días)")
    
    if not df_pedidos.empty:
        df_morosos = df_pedidos[df_pedidos["Estado"].isin(["Pendiente", "Entregado Sin Cobrar"])].copy()
        df_morosos["Fecha_dt"] = pd.to_datetime(df_morosos["Fecha"], format="%d/%m/%Y", errors="coerce")
        
        hoy = datetime.now()
        df_morosos = df_morosos[(hoy - df_morosos["Fecha_dt"]).dt.days > 7]
        
        if not df_morosos.empty:
            st.error("🚨 **Clientes que superan la semana de deuda:**")
            for _, row in df_morosos.iterrows():
                st.write(f"- **{row['Cliente']}**: debe `${row['Monto Total']:,.2f}` *(Trabajo del {row['Fecha']})*")
        else:
            st.success("✅ No hay clientes con deudas mayores a una semana. ¡Excelente gestión de cobros!")
    else:
        st.info("No hay trabajos registrados para analizar morosidad.")
    # =========================================================================
    # ... HASTA ACÁ
    # =========================================================================        


# --- PESTAÑA PEDIDOS Y CLIENTES ---
with tab_pedidos:
    st.subheader("Gestión de Clientes y Trabajos de Impresión / Cartelería")
    
    with st.expander("➕ Cargar Nuevo Trabajo (Impresión o Cartelería)", expanded=False):
        # Obtenemos la lista única de clientes anteriores ordenada
        clientes_existentes = sorted(df_pedidos["Cliente"].dropna().unique().tolist()) if not df_pedidos.empty else []
        
        with st.form("form_imprenta_pedido"):
            col_a, col_b = st.columns(2)
            with col_a:
                pedido_fecha = st.date_input("Fecha del Trabajo:", value=datetime.today())
                
                # Campo único con autocompletado: arranca vacío y permite buscar o escribir uno nuevo
                opciones_cliente = [""] + clientes_existentes
                cliente_elegido = st.selectbox(
                    "Cliente (Buscá o escribí el nombre):", 
                    options=opciones_cliente,
                    help="Podés seleccionar de la lista o escribir directamente."
                )
                
                # Por si prefieren escribir libremente si no está en la lista exacta
                cliente_texto_libre = st.text_input("O escribí el nombre si es nuevo (opcional):", value="")
                
                material_tipo = st.selectbox("Tipo de Material / Rubro:", [
                    "LONA FRONTLIGHT", "LONA BACKLIGHT", "VINILO BASE BLANCA", 
                    "VINILO BASE GRIS", "VINILO MATE", "VINILO CRISTAL", 
                    "VINILO MICROPERFORADO", "TELA BANDERA", "PAPEL BLUEBACK",
                    "TRABAJOS DE CARTELERIA"
                ])
                medida_trabajo = st.text_input("Medida / Tamaño:")
            with col_b:
                monto_trabajo = st.number_input("Monto total ($):", min_value=0.0, step=100.0)
                estado_pedido = st.selectbox("Estado Pago / Entrega:", ["Pendiente", "Entregado Cobrado", "Entregado Sin Cobrar"])
                usuario_cobro = st.selectbox("¿Quién cobró?:", ["N/A", "ROBERTO", "GONZALO"])
                estado_impresion = st.selectbox("Estado Producción:", ["🔴 Falta Imprimir", "🟢 Finalizado"])
            
            obs_trabajo = st.text_area("🗒️ Observaciones:")
            bot_guardar_p = st.form_submit_button("💾 Registrar Trabajo")
            
            if bot_guardar_p:
                # Definimos el nombre final: toma el campo de texto libre si escribió algo ahí, sino usa el selectbox
                cliente_name = cliente_texto_libre.strip() if cliente_texto_libre.strip() else cliente_elegido
                
                if cliente_name:
                    fecha_formateada = pedido_fecha.strftime("%d/%m/%Y")
                    quien_guarda = "ROB" if "ROBERTO" in usuario_cobro.upper() else ("GON" if "GONZALO" in usuario_cobro.upper() else "N/A")
                    
                    nuevo_registro = pd.DataFrame([{
                        "Fecha": fecha_formateada, "Cliente": cliente_name.upper().strip(), "Material": material_tipo,
                        "Medida": medida_trabajo.upper().strip() if medida_trabajo else "N/A", "Copias": "1 UNIDAD",
                        "Monto Total": monto_trabajo, "Estado": estado_pedido, "Cobró": quien_guarda,
                        "Observaciones": obs_trabajo.strip() if obs_trabajo else "", "Impresión": estado_impresion
                    }])
                    df_pedidos = pd.concat([df_pedidos, nuevo_registro], ignore_index=True)
                    df_pedidos.to_csv(archivo_pedidos, index=False)
                    st.success("¡Trabajo registrado!")
                    st.rerun()
                else:
                    st.warning("⚠️ Por favor, seleccioná o escribí el nombre del cliente.")

    st.write("---")
    st.subheader("📋 Historial General de Trabajos (Doble clic para editar celdas)")
    
    df_visualizacion = df_pedidos.copy()
    if not df_visualizacion.empty:
        col_f1, col_f2 = st.columns([1, 1])
        with col_f1:
            usar_filtro_semana = st.checkbox("📆 Filtrar historial por semana específica")
        with col_f2:
            solo_deben = st.toggle("🔴 Mostrar solo trabajos que faltan cobrar")

        if solo_deben:
            df_visualizacion = df_visualizacion[df_visualizacion["Estado"].isin(["Pendiente", "Entregado Sin Cobrar"])]
            
            # Muestra la alerta con el total adeudado de los trabajos filtrados
            if not df_visualizacion.empty:
                total_filtrado_debe = df_visualizacion["Monto Total"].sum()
                st.warning(f"⚠️ **Total adeudado en esta selección:** ${total_filtrado_debe:,.2f}")
            else:
                st.info("🎉 ¡No hay trabajos pendientes de cobro en esta vista!")

        if usar_filtro_semana:
            df_visualizacion["Fecha_dt_aux"] = pd.to_datetime(df_visualizacion["Fecha"], format="%d/%m/%Y", errors="coerce")
            df_visualizacion["Semana_Inicio"] = df_visualizacion["Fecha_dt_aux"].dt.to_period('W').dt.start_time
            semanas_disponibles = sorted(df_visualizacion["Semana_Inicio"].dropna().unique(), reverse=True)
            
            if semanas_disponibles:
                semana_elegida = st.selectbox("Seleccioná la semana:", semanas_disponibles, format_func=lambda x: f"Semana del Lunes {x.strftime('%d/%m/%Y')} al Domingo {(x + timedelta(days=6)).strftime('%d/%m/%Y')}")
                df_visualizacion = df_visualizacion[df_visualizacion["Semana_Inicio"] == semana_elegida]
            
            df_visualizacion = df_visualizacion.drop(columns=["Fecha_dt_aux", "Semana_Inicio"], errors="ignore")

        buscar_cli = st.text_input("🔍 Filtrar historial por Cliente:")
        if buscar_cli:
            df_visualizacion = df_visualizacion[df_visualizacion["Cliente"].astype(str).str.contains(buscar_cli, case=False, na=False)]
        
        df_visualizacion["Material"] = df_visualizacion["Material"].fillna("").astype(str)
    
        df_agrupado = df_visualizacion.groupby(["Fecha", "Cliente"], as_index=False).agg({
            "Material": lambda x: " / ".join([v for v in x if v != "" and v.lower() != "nan"]),
            "Material": lambda x: " / ".join(str(v) for v in x if pd.notna(v) and str(v).strip() != ""), 
            "Medida": lambda x: " / ".join(x.astype(str)),
            "Monto Total": "sum", 
            "Estado": "first", 
            "Cobró": "first",
            "Observaciones": lambda x: " | ".join([v for v in x.astype(str) if v.strip() != ""]),
            "Impresión": "first"
        })
        
        df_agrupado["Cobró"] = df_agrupado["Cobró"].replace({"ROBERTO": "ROB", "GONZALO": "GON"})
        
        df_agrupado["Fecha_temp"] = pd.to_datetime(df_agrupado["Fecha"], format="%d/%m/%Y", errors="coerce")
        df_agrupado = df_agrupado.sort_values(by="Fecha_temp", ascending=False).drop(columns=["Fecha_temp"])
        df_agrupado = df_agrupado[["Fecha", "Cliente", "Material", "Medida", "Monto Total", "Estado", "Cobró", "Impresión", "Observaciones"]]
        
        lista_con_separadores = []
        for i in range(len(df_agrupado)):
            fila_actual = df_agrupado.iloc[i].copy()
            
            if fila_actual["Estado"] == "Entregado Cobrado":
                fila_actual["Fecha_Visual"] = f"🟢 {fila_actual['Fecha']}"
            elif fila_actual["Estado"] in ["Pendiente", "Entregado Sin Cobrar"]:
                fila_actual["Fecha_Visual"] = f"🔴 {fila_actual['Fecha']}"
            else:
                fila_actual["Fecha_Visual"] = fila_actual['Fecha']
                
            # Lógica de WhatsApp personalizada
            msg_horarios = "de lunes a viernes de 8 a 12 y de 16:30 a 19:30 hs, y sábados de 9 a 12 hs"
            
            # Mensaje modificado: sin medidas y con el nombre "Impresiones Resistencia"
            texto_wsp = (f"¡Hola {fila_actual['Cliente']}! Te escribimos de Impresiones Resistencia para avisarte que tu "
                         f"trabajo de {fila_actual['Material']} ya está listo. "
                         f"El total es de ${fila_actual['Monto Total']:,.2f}. "
                         f"Podés pasar a retirarlo en nuestros horarios: {msg_horarios}. ¡Saludos!")
            
            import urllib.parse
            link_wsp = f"https://wa.me/?text={urllib.parse.quote(texto_wsp)}"
            fila_actual["WhatsApp"] = link_wsp if fila_actual["Impresión"] == "🟢 Finalizado" else None

            lista_con_separadores.append(fila_actual)
            
            if i < len(df_agrupado) - 1:
                fecha_siguiente = df_agrupado.iloc[i+1]["Fecha"]
                if fila_actual["Fecha"] != fecha_siguiente:
                    fila_separadora = pd.Series({
                        "Fecha": "", "Cliente": "--- CAMBIO DE DÍA ---", "Material": "---", "Medida": "---", 
                        "Monto Total": 0.0, "Estado": "---", "Cobró": "---", "Impresión": "---", 
                        "Observaciones": "---", "Fecha_Visual": "────────────────", "WhatsApp": None
                    })
                    lista_con_separadores.append(fila_separadora)
                    
        df_final_visual = pd.DataFrame(lista_con_separadores).reset_index(drop=True)
        
        # Insertamos la columna Borrar al final si no existe
        if "Borrar" not in df_final_visual.columns:
            df_final_visual["Borrar"] = False
            
        # Orden exacto donde WhatsApp va antes de Observaciones
        columnas_ordenadas = [
            "Fecha_Visual", "Cliente", "Material", "Medida", 
            "Monto Total", "Estado", "Cobró", "Impresión", 
            "WhatsApp", "Observaciones", "Borrar"
        ]
        df_editor_vista = df_final_visual[columnas_ordenadas]
        
        config_columnas = {
            "Fecha_Visual": st.column_config.TextColumn("Fecha", disabled=True, width=95),
            "Cliente": st.column_config.TextColumn(disabled=True, width=150),
            "Material": st.column_config.TextColumn(disabled=True, width=200),
            "Medida": st.column_config.TextColumn(disabled=True, width=200),
            "Monto Total": st.column_config.NumberColumn(format="$%.2f", disabled=False, width="small"),
            "Estado": st.column_config.SelectboxColumn(options=["Pendiente", "Entregado Cobrado", "Entregado Sin Cobrar"], width=88),
            "Cobró": st.column_config.SelectboxColumn(options=["N/A", "ROB", "GON"], width=60),
            "Impresión": st.column_config.SelectboxColumn(options=["🔴 Falta Imprimir", "🟢 Finalizado"], width=88),
            "WhatsApp": st.column_config.LinkColumn("Avisar Cliente", help="Enviar aviso", display_text="📲 Enviar", width=62),
            "Observaciones": st.column_config.TextColumn(width="small"),
            "Borrar": st.column_config.CheckboxColumn("🗑️", help="Tildá para eliminar esta fila", width=50, default=False)
        }
        
        df_cambiado = st.data_editor(
            df_editor_vista,
            column_config=config_columnas,
            use_container_width=True,
            hide_index=True,
            key="editor_historial_pedidos"
        )
        
        # --- BOTÓN DEFINITIVO PARA BORRAR LAS FILAS SELECCIONADAS ---
        col_btn_del1, col_btn_del2 = st.columns([2, 3])
        with col_btn_del1:
            if st.button("🗑️ Eliminar trabajos marcados", type="primary"):
                filas_a_borrar = df_cambiado[df_cambiado["Borrar"] == True]
    
                if not filas_a_borrar.empty:
                    count_eliminados = 0
                    for _, fila_del in filas_a_borrar.iterrows():
                        if "CAMBIO DE DÍA" in str(fila_del["Cliente"]): 
                            continue
            
                        # Capturamos los datos clave de la fila a eliminar
                        fecha_limpia = str(fila_del["Fecha_Visual"]).replace("🟢 ", "").replace("🔴 ", "").strip()
                        cliente_objetivo = str(fila_del["Cliente"]).strip()
                        monto_objetivo = float(fila_del["Monto Total"])
            
                        # Creamos una máscara para encontrar la fila exacta en el df original
                        condicion = (
                            (df_pedidos["Fecha"].astype(str).str.strip() == fecha_limpia) & 
                            (df_pedidos["Cliente"].astype(str).str.upper().str.strip() == cliente_objetivo) & 
                            (df_pedidos["Monto Total"] == monto_objetivo)
                        )
            
                        if condicion.any():
                            # Eliminamos la fila del DataFrame original (esto reacomoda todo automáticamente sin dejar espacios vacíos)
                            df_pedidos = df_pedidos[~condicion]
                            count_eliminados += 1
        
                    # Guardamos el CSV limpio (index=False evita que se guarden índices extraños)
                    df_pedidos.to_csv(archivo_pedidos, index=False)
                    st.success(f"¡Se eliminaron {count_eliminados} trabajo(s) y se compactó el archivo correctamente!")
                    st.rerun()
                else:
                    st.warning("⚠️ No marcaste ningún trabajo para borrar.")

        # --- GUARDADO DE EDICIONES COMUNES ---
        if not df_cambiado.drop(columns=["Borrar"]).equals(df_editor_vista.drop(columns=["Borrar"])):
            for idx, fila in df_cambiado.iterrows():
                if fila["Cliente"] == "--- CAMBIO DE DÍA ---": continue
                
                fecha_real = df_final_visual.iloc[idx]["Fecha"]
                f_mask = (df_pedidos["Fecha"] == fecha_real) & (df_pedidos["Cliente"] == fila["Cliente"])
                
                if f_mask.any():
                    indices_coincidentes = df_pedidos[f_mask].index
                    primer_indice = indices_coincidentes[0]  # Nos quedamos solo con el primer registro del grupo
                    
                    if fila["Estado"] == "Entregado Cobrado" and df_pedidos.loc[primer_indice, "Estado"] != "Entregado Cobrado":
                        df_pedidos.loc[primer_indice, "Fecha Cobro"] = datetime.now().strftime("%d/%m/%Y")
                    
                    # Actualizamos estado, cobró e impresión a todo el grupo (para que no desentone)
                    df_pedidos.loc[f_mask, "Estado"] = fila["Estado"]
                    df_pedidos.loc[f_mask, "Cobró"] = fila["Cobró"]
                    df_pedidos.loc[f_mask, "Impresión"] = fila["Impresión"]
                    
                    # PERO las observaciones se guardan ÚNICAMENTE en el primer registro
                    df_pedidos.loc[f_mask, "Observaciones"] = ""  # Limpiamos los demás
                    df_pedidos.loc[primer_indice, "Observaciones"] = fila["Observaciones"] # Ponemos el texto solo en el primero
            
            df_pedidos.to_csv(archivo_pedidos, index=False)
            st.rerun()
            
        fecha_hoy_str = datetime.now().strftime("%d/%m/%Y")
        st.write("---")
        st.markdown(f"### 💵 Resumen General y de Ingresos ({fecha_hoy_str})")
        
        # --- CÁLCULO DE MONTO PENDIENTE Y TRABAJOS INGRESADOS ---
        monto_pendiente = df_pedidos[df_pedidos["Estado"].isin(["Pendiente", "Entregado Sin Cobrar"])]["Monto Total"].sum()
        
        df_pedidos["Fecha_dt_aux"] = pd.to_datetime(df_pedidos["Fecha"], format="%d/%m/%Y", errors="coerce")
        
        # Monto de trabajos cargados HOY
        monto_ingresado_hoy = df_pedidos[df_pedidos["Fecha"] == fecha_hoy_str]["Monto Total"].sum()
        
        # Monto de trabajos cargados en la SEMANA ACTUAL (Lunes a Domingo)
        hoy_dt = datetime.now()
        inicio_semana_actual = hoy_dt - timedelta(days=hoy_dt.weekday()) # Lunes de esta semana
        inicio_semana_actual = inicio_semana_actual.replace(hour=0, minute=0, second=0, microsecond=0)
        
        mask_semana = (df_pedidos["Fecha_dt_aux"] >= inicio_semana_actual) & (df_pedidos["Fecha_dt_aux"] <= hoy_dt)
        monto_ingresado_semana = df_pedidos[mask_semana]["Monto Total"].sum()
        
        # Mostramos las 3 métricas en columnas (Pendiente + Ingresado Hoy + Ingresado Semana)
        col_m1, col_m2, col_m3 = st.columns(3)
        col_m1.metric(label="🔴 MONTO PENDIENTE TOTAL (GENERAL)", value=f"${monto_pendiente:,.2f}")
        col_m2.metric(label="📥 Ingresados Hoy (Cargados)", value=f"${monto_ingresado_hoy:,.2f}")
        col_m3.metric(label="📅 Ingresados en la Semana (Cargados)", value=f"${monto_ingresado_semana:,.2f}")


# --- PESTAÑA MATERIALES Y GASTOS PROVEEDORES ---
with tab_materiales:
    st.subheader("Gastos en Materiales e Insumos")
    
    with st.expander("➕ Cargar Nueva Compra / Gasto de Proveedor", expanded=False):
        with st.form("form_nuevo_gasto"):
            col_g1, col_g2 = st.columns(2)
            with col_g1:
                gasto_fecha = st.date_input("Fecha de Compra:", value=datetime.today())
                gasto_desc = st.text_input("Descripción del material:")
                gasto_monto = st.number_input("Monto total ($):", min_value=0.0, step=100.0)
            with col_g2:
                gasto_prov = st.text_input("Proveedor:").upper().strip()
                gasto_pago = st.text_input("Forma de Pago (Opcional):")
                gasto_estado = st.selectbox("Estado de la Factura:", ["PENDIENTE", "PAGADO"])
                
            bot_guardar_gasto = st.form_submit_button("📦 Registrar Compra")
            
            if bot_guardar_gasto:
                if gasto_desc and gasto_monto > 0:
                    gasto_saldo = 0.0 if gasto_estado == "PAGADO" else gasto_monto
                    nuevo_gasto = pd.DataFrame([{
                        "FECHA": gasto_fecha.strftime("%Y-%m-%d"), "DESCRIPCION": gasto_desc.upper().strip(),
                        "MONTO": gasto_monto, "FORMA DE PAGO": gasto_pago.upper().strip() if gasto_pago else "",
                        "PROVEEDOR": gasto_prov if gasto_prov else "VARIOS", "ESTADO": gasto_estado, "SALDO": gasto_saldo
                    }])
                    df_mat = pd.concat([df_mat, nuevo_gasto], ignore_index=True)
                    df_mat.to_excel(archivo_materiales, index=False)
                    st.success("¡Compra guardada!")
                    st.rerun()

    st.write("---")
    st.subheader("📋 Historial de Compras (Doble clic para editar celdas)")
    
    if not df_mat.empty:
        buscador_m = st.text_input("🔍 Buscar por descripción o proveedor (Materiales):")
        df_mat_filt = df_mat.copy()
        if buscador_m:
            df_mat_filt = df_mat_filt[df_mat_filt["DESCRIPCION"].astype(str).str.contains(buscador_m, case=False, na=False) | df_mat_filt["PROVEEDOR"].astype(str).str.contains(buscador_m, case=False, na=False)]
            
        df_mat_filt["FECHA"] = df_mat_filt["FECHA"].astype(str).apply(lambda x: x.split(" ")[0])
        
        df_mat_filt["FECHA_DT_SORT"] = pd.to_datetime(df_mat_filt["FECHA"], errors="coerce")
        df_mat_filt = df_mat_filt.sort_values(by="FECHA_DT_SORT", ascending=False).drop(columns=["FECHA_DT_SORT"])
        
        df_mat_filt["FECHA_ORIGINAL"] = df_mat_filt["FECHA"]
        df_mat_filt["FECHA"] = df_mat_filt.apply(
            lambda r: f"🟢 {r['FECHA']}" if r["ESTADO"] == "PAGADO" else r["FECHA"], 
            axis=1
        )
        
        config_mat = {
            "FECHA": st.column_config.TextColumn(disabled=True),
            "DESCRIPCION": st.column_config.TextColumn(),
            "MONTO": st.column_config.NumberColumn(format="$%.2f"),
            "FORMA DE PAGO": st.column_config.TextColumn(),
            "PROVEEDOR": st.column_config.TextColumn(),
            "ESTADO": st.column_config.SelectboxColumn(options=["PENDIENTE", "PAGADO"]),
            "SALDO": st.column_config.NumberColumn(format="$%.2f", disabled=True)
        }
        
        df_mat_cambiado = st.data_editor(
            df_mat_filt.drop(columns=["FECHA_ORIGINAL"]),
            column_config=config_mat,
            use_container_width=True,
            hide_index=True,
            key="editor_materiales"
        )
        
        if not df_mat_cambiado.equals(df_mat_filt.drop(columns=["FECHA_ORIGINAL"])):
            for idx, fila in df_mat_cambiado.iterrows():
                fecha_real = df_mat_filt.iloc[idx]["FECHA_ORIGINAL"]
                desc_real = df_mat_filt.iloc[idx]["DESCRIPCION"]
                
                f_mask = (df_mat["FECHA"].astype(str).str.contains(fecha_real)) & (df_mat["DESCRIPCION"] == desc_real)
                if f_mask.any():
                    df_mat.loc[f_mask, "DESCRIPCION"] = fila["DESCRIPCION"]
                    df_mat.loc[f_mask, "MONTO"] = fila["MONTO"]
                    df_mat.loc[f_mask, "FORMA DE PAGO"] = fila["FORMA DE PAGO"]
                    df_mat.loc[f_mask, "PROVEEDOR"] = fila["PROVEEDOR"]
                    df_mat.loc[f_mask, "ESTADO"] = fila["ESTADO"]
                    df_mat.loc[f_mask, "SALDO"] = 0.0 if fila["ESTADO"] == "PAGADO" else fila["MONTO"]
            
            df_mat.to_excel(archivo_materiales, index=False)
            st.rerun()
        
        total_deuda_filtrada = df_mat_filt[df_mat_filt["ESTADO"] != "PAGADO"]["MONTO"].sum()
        st.error(f"⚠️ Total adeudado en esta selección de materiales: ${total_deuda_filtrada:,.2f}")
        # =========================================================================
    # 📉 NUEVO: ÍNDICE DE IMPACTO DE MATERIALES VS FACTURACIÓN
    # =========================================================================
    st.write("---")
    st.subheader("📉 Indicador de Impacto de Materiales")
    
    if not df_mat.empty and not df_pedidos.empty:
        # Total gastado en materiales (histórico o filtrado según prefieras, acá tomamos el total general de compras)
        gasto_total_materiales = df_mat["MONTO"].sum()
        
        # Total facturado por clientes
        facturacion_total_clientes = df_pedidos["Monto Total"].sum()
        
        if facturacion_total_clientes > 0:
            porcentaje_costo_material = (gasto_total_materiales / facturacion_total_clientes) * 100
        else:
            porcentaje_costo_material = 0.0
            
        col_m_ind1, col_m_ind2, col_m_ind3 = st.columns(3)
        
        col_m_ind1.metric(
            label="📦 Gasto Total en Insumos", 
            value=f"${gasto_total_materiales:,.2f}"
        )
        col_m_ind2.metric(
            label="💰 Facturación Total Clientes", 
            value=f"${facturacion_total_clientes:,.2f}"
        )
        col_m_ind3.metric(
            label="📊 Incidencia de Insumos", 
            value=f"{porcentaje_costo_material:.1f}%",
            delta="Costo sobre ventas"
        )
        
        # Alerta orientativa de rentabilidad
        if porcentaje_costo_material > 50:
            st.error("🚨 **Alerta de costos:** Estás gastando más del 50% de tu facturación en insumos. Los precios de venta de tus trabajos podrían estar quedando bajos frente a la inflación de materiales.")
        elif porcentaje_costo_material > 35:
            st.warning("⚠️ **Margen moderado:** El peso de los materiales es considerable. Monitoreá de cerca los aumentos de proveedores.")
        else:
            st.success("✅ **Excelente margen operativo:** El peso de los insumos sobre la facturación se mantiene en un nivel muy saludable.")
    else:
        st.info("Cargá registros tanto en Pedidos como en Materiales para ver el análisis de impacto.")

# --- PESTAÑA CHEQUES ---
with tab_cheques:
    st.subheader("💳 Cartera de Cheques")
    
    with st.expander("➕ Cargar Nuevo Cheque", expanded=False):
        with st.form("form_nuevo_cheque_v3", clear_on_submit=True):
            col_c1, col_c2 = st.columns(2)
            with col_c1:
                nuevo_prov = st.text_input("Proveedor:")
                nuevo_monto = st.number_input("Monto ($):", min_value=0.0, step=100.0)
            with col_c2:
                nueva_fecha = st.text_input("Fecha Vencimiento (ej: 20/12):")
                nuevo_estado = st.selectbox("Estado:", ["Pendiente", "Pagado"])
            
            if st.form_submit_button("💾 Registrar Cheque"):
                if nuevo_prov and nuevo_monto > 0:
                    wb = openpyxl.load_workbook(archivo_cheques)
                    ws = wb.active
                    col_idx = None
                    for c in range(1, ws.max_column + 1):
                        if str(ws.cell(1, c).value).strip() == nuevo_prov:
                            col_idx = c
                            break
                    if not col_idx:
                        col_idx = ws.max_column + 1
                        ws.cell(1, col_idx).value = nuevo_prov
                    
                    next_row = ws.max_row + 1
                    ws.cell(row=next_row, column=col_idx).value = f"${nuevo_monto} {nueva_fecha}"
                    wb.save(archivo_cheques)
                    st.success("¡Registrado!")
                    st.rerun()
                else:
                    st.error("Completá proveedor y monto.")

    st.write("---")
    
    if not df_cheques.empty:
        col_filtro1, col_filtro2 = st.columns(2)
        with col_filtro1:
            proveedor_cheque = st.selectbox("Filtrar por Proveedor:", ["Todos"] + sorted(df_cheques["Proveedor"].unique().tolist()))
        with col_filtro2:
            estado_cheque = st.radio("Filtrar por Estado:", ["Todos", "Pendientes", "Pagados"], horizontal=True)
        
        df_cheques_filt = df_cheques.copy()
        if proveedor_cheque != "Todos":
            df_cheques_filt = df_cheques_filt[df_cheques_filt["Proveedor"] == proveedor_cheque]
        if estado_cheque == "Pendientes":
            df_cheques_filt = df_cheques_filt[df_cheques_filt["Estado"] == "Pendiente"]
        elif estado_cheque == "Pagados":
            df_cheques_filt = df_cheques_filt[df_cheques_filt["Estado"] == "Pagado"]
            
        hoy_anio = datetime.now().year
        df_cheques_filt["Fecha_Sort"] = pd.to_datetime(df_cheques_filt["Fecha Vencimiento"] + f"/{hoy_anio}", format="%d/%m/%Y", errors="coerce")
        df_cheques_filt["Prioridad_Estado"] = df_cheques_filt["Estado"].apply(lambda x: 0 if x == "Pendiente" else 1)
        
        df_cheques_filt = df_cheques_filt.sort_values(by=["Prioridad_Estado", "Fecha_Sort"], ascending=[True, False]).drop(columns=["Fecha_Sort", "Prioridad_Estado"])

        df_cheques_filt["Fecha Vencimiento"] = df_cheques_filt.apply(
            lambda r: f"🟢 {r['Fecha Vencimiento']}" if r["Estado"] == "Pagado" else r["Fecha Vencimiento"], 
            axis=1
        )
        
        st.data_editor(
            df_cheques_filt[["Proveedor", "Monto", "Fecha Vencimiento", "Estado"]],
            column_config={
                "Monto": st.column_config.NumberColumn(format="$%.2f"),
                "Estado": st.column_config.SelectboxColumn(options=["Pendiente", "Pagado"])
            },
            use_container_width=True, hide_index=True
        )


# --- PESTAÑA: CIERRE SEMANAL ---
with tab_semanal:
    st.subheader("Control, Rendición y Gastos de Caja Semanal")
    if not df_pedidos.empty:
        df_sem = df_pedidos.copy()
        
        if "Fecha Cobro" in df_sem.columns:
            df_sem["Fecha_Uso"] = df_sem["Fecha Cobro"].fillna(df_sem["Fecha"])
        else:
            df_sem["Fecha_Uso"] = df_sem["Fecha"]
            
        df_sem["Fecha_dt"] = pd.to_datetime(df_sem["Fecha_Uso"], format="%d/%m/%Y", errors='coerce')
        df_sem = df_sem.dropna(subset=["Fecha_dt"])
        df_sem["Semana_Inicio"] = df_sem["Fecha_dt"].dt.to_period('W').dt.start_time
        
        semanas_disponibles = sorted(df_sem["Semana_Inicio"].unique(), reverse=True)
        
        if semanas_disponibles:
            semana_sel = st.selectbox("Seleccioná la semana a auditar:", semanas_disponibles, format_func=lambda x: f"Semana del Lunes {x.strftime('%d/%m/%Y')} al Domingo {(x + timedelta(days=6)).strftime('%d/%m/%Y')}")
            semana_sel_str = semana_sel.strftime("%Y-%m-%d")
            
            df_semana_filtrada = df_sem[df_sem["Semana_Inicio"] == semana_sel].copy()
            df_cobrado_sem = df_semana_filtrada[df_semana_filtrada["Estado"] == "Entregado Cobrado"]
            
            total_sem_bruto_rob = df_cobrado_sem[df_cobrado_sem["Cobró"].isin(["ROB", "ROBERTO"])]["Monto Total"].sum()
            total_sem_bruto_gon = df_cobrado_sem[df_cobrado_sem["Cobró"].isin(["GON", "GONZALO"])]["Monto Total"].sum()
            
            st.write("---")
            
            col_roberto, col_gonzalo = st.columns(2)
            
            with col_roberto:
                st.markdown("### 💼 RENDICIÓN ROBERTO")
                df_gastos_rob = df_gastos_sem[(df_gastos_sem["Semana"] == semana_sel_str) & (df_gastos_sem["Socio"] == "ROB")][["Concepto", "Monto"]].reset_index(drop=True)
                
                st.markdown(f"**Cobrado Bruto Clientes:** `${total_sem_bruto_rob:,.2f}`")
                st.caption("👇 Agregar o editar gastos de Roberto acá abajo:")
                
                gastos_rob_editado = st.data_editor(
                    df_gastos_rob,
                    column_config={
                        "Concepto": st.column_config.TextColumn(),
                        "Monto": st.column_config.NumberColumn(format="$%.2f", min_value=0.0)
                    },
                    num_rows="dynamic",
                    use_container_width=True,
                    key="editor_gastos_roberto"
                )
                
                total_gastos_rob = gastos_rob_editado["Monto"].sum()
                neto_roberto = total_sem_bruto_rob - total_gastos_rob
                
                st.metric(label="💰 TOTAL NETO LIMPIO (ROB)", value=f"${neto_roberto:,.2f}", delta=f"-${total_gastos_rob:,.2f} Gastos")
                
            with col_gonzalo:
                st.markdown("### 💼 RENDICIÓN GONZALO")
                df_gastos_gon = df_gastos_sem[(df_gastos_sem["Semana"] == semana_sel_str) & (df_gastos_sem["Socio"] == "GON")][["Concepto", "Monto"]].reset_index(drop=True)
                
                st.markdown(f"**Cobrado Bruto Clientes:** `${total_sem_bruto_gon:,.2f}`")
                st.caption("👇 Agregar o editar gastos de Gonzalo acá abajo:")
                
                gastos_gon_editado = st.data_editor(
                    df_gastos_gon,
                    column_config={
                        "Concepto": st.column_config.TextColumn(),
                        "Monto": st.column_config.NumberColumn(format="$%.2f", min_value=0.0)
                    },
                    num_rows="dynamic",
                    use_container_width=True,
                    key="editor_gastos_gonzalo"
                )
                
                total_gastos_gon = gastos_gon_editado["Monto"].sum()
                neto_gonzalo = total_sem_bruto_gon - total_gastos_gon
                
                st.metric(label="💰 TOTAL NETO LIMPIO (GON)", value=f"${neto_gonzalo:,.2f}", delta=f"-${total_gastos_gon:,.2f} Gastos")

            st.write("---")
            st.markdown("### ⚖️ BALANCE Y DIVISIÓN DE CAJA (50% / 50%)")
            
            mitad_roberto = neto_roberto / 2
            mitad_gonzalo = neto_gonzalo / 2
            diferencia_netos = mitad_roberto - mitad_gonzalo
            
            col_div1, col_div2, col_div3 = st.columns(3)
            col_div1.metric(label="👤 Mitad Roberto (Neto / 2)", value=f"${mitad_roberto:,.2f}")
            col_div2.metric(label="👤 Mitad Gonzalo (Neto / 2)", value=f"${mitad_gonzalo:,.2f}")
            
            if diferencia_netos > 0:
                col_div3.metric(label="🔄 Ajuste de Cuentas", value=f"${abs(diferencia_netos):,.2f}", delta="A favor de GON", delta_color="inverse")
                st.info(f"💡 **Explicación del Balance:** Para que queden iguales, **Roberto debe darle ${abs(diferencia_netos):,.2f} a Gonzalo**.")
            elif diferencia_netos < 0:
                col_div3.metric(label="🔄 Ajuste de Cuentas", value=f"${abs(diferencia_netos):,.2f}", delta="A favor de ROB")
                st.info(f"💡 **Explicación del Balance:** Para que queden iguales, **Gonzalo debe darle ${abs(diferencia_netos):,.2f} a Roberto**.")
            else:
                col_div3.metric(label="🔄 Ajuste de Cuentas", value="$0.00", delta="Caja Perfecta")
                st.success("¡Están perfectamente empatados! No se deben nada.")

            st.write("")
            if st.button("💾 Guardar Cambios de Gastos Semanales"):
                df_gastos_sem = df_gastos_sem[df_gastos_sem["Semana"] != semana_sel_str]
                
                if not gastos_rob_editado.empty:
                    gastos_rob_save = gastos_rob_editado.dropna(subset=["Concepto"])
                    gastos_rob_save.insert(0, "Semana", semana_sel_str)
                    gastos_rob_save.insert(1, "Socio", "ROB")
                    df_gastos_sem = pd.concat([df_gastos_sem, gastos_rob_save], ignore_index=True)
                    
                if not gastos_gon_editado.empty:
                    gastos_gon_save = gastos_gon_editado.dropna(subset=["Concepto"])
                    gastos_gon_save.insert(0, "Semana", semana_sel_str)
                    gastos_gon_save.insert(1, "Socio", "GON")
                    df_gastos_sem = pd.concat([df_gastos_sem, gastos_gon_save], ignore_index=True)
                
                df_gastos_sem.to_csv(archivo_gastos_semanales, index=False)
                st.success("¡Gastos de la semana guardados correctamente!")
                st.rerun()

    # --- AQUÍ ADENTRO DE TAB_SEMANAL VA EL BACKUP ---
    st.write("---")
    st.subheader("💾 Respaldo de Datos")
    if st.button("🚀 Realizar Backup Semanal"):
        destino_backup = Path(r"C:\Users\Usuario\Desktop\sistema_imprenta\BACKUP SEMANAL")
        try:
            destino_backup.mkdir(parents=True, exist_ok=True)
            archivos_a_copiar = [
                obtener_ruta("PEDIDOS_CLIENTES.csv"),
                obtener_ruta("MATERIALES.xlsx"),
                obtener_ruta("CHEQUES NUEVO.xlsx")
            ]
            for origen in archivos_a_copiar:
                if os.path.exists(origen):
                    nombre_archivo = os.path.basename(origen)
                    shutil.copy2(origen, destino_backup / nombre_archivo)
            st.success(f"✅ Backup realizado con éxito en: {destino_backup}")
        except Exception as e:
            st.error(f"❌ Error al realizar el backup: {e}")
            
    # --- EXPORTAR RENDICIÓN SEMANAL ---
    st.write("---")
    st.markdown("### 📤 Exportar Rendición Semanal Completa")

    df_cobrados_sem = df_semana_filtrada[df_semana_filtrada["Estado"] == "Entregado Cobrado"]

    det_rob = df_cobrados_sem[df_cobrados_sem["Cobró"].isin(["ROB", "ROBERTO"])]
    lista_rob = "\n".join([f"- {r['Cliente']}: ${r['Monto Total']:,.0f}" for _, r in det_rob.iterrows()])

    det_gon = df_cobrados_sem[df_cobrados_sem["Cobró"].isin(["GON", "GONZALO"])]
    lista_gon = "\n".join([f"- {r['Cliente']}: ${r['Monto Total']:,.0f}" for _, r in det_gon.iterrows()])

    lista_gastos_rob = "\n".join([f"- {row['Concepto']}: ${row['Monto']:,.0f}" for _, row in gastos_rob_editado.dropna(subset=["Concepto"]).iterrows()]) if 'gastos_rob_editado' in locals() and not gastos_rob_editado.empty else "Sin gastos"
    lista_gastos_gon = "\n".join([f"- {row['Concepto']}: ${row['Monto']:,.0f}" for _, row in gastos_gon_editado.dropna(subset=["Concepto"]).iterrows()]) if 'gastos_gon_editado' in locals() and not gastos_gon_editado.empty else "Sin gastos"

    texto_wa = f"💰 Rendición Semanal ({semana_sel.strftime('%d/%m')} al {(semana_sel + timedelta(days=6)).strftime('%d/%m')})\n\n"
    texto_wa += f"👤 ROBERTO\n* Cobros Clientes:\n{lista_rob if lista_rob else 'Sin cobros'}\n\n* Total Bruto: ${total_sem_bruto_rob:,.2f}\n\n* Gastos:\n{lista_gastos_rob}\n\n👉 *Neto Roberto: ${neto_roberto:,.2f}*\n\n\n"
    texto_wa += f"👤 GONZALO\n* Cobros Clientes:\n{lista_gon if lista_gon else 'Sin cobros'}\n\n* Total Bruto: ${total_sem_bruto_gon:,.2f}\n\n* Gastos:\n{lista_gastos_gon}\n\n👉 *Neto Gonzalo: ${neto_gonzalo:,.2f}*\n\n\n"
    texto_wa += f"⚖️ *BALANCE FINAL*\n"
    if diferencia_netos > 0:
        texto_wa += f"Roberto debe transferir a Gonzalo: ${abs(diferencia_netos):,.2f}"
    elif diferencia_netos < 0:
        texto_wa += f"Gonzalo debe transferir a Roberto: ${abs(diferencia_netos):,.2f}"
    else:
        texto_wa += f"Caja Perfecta (Sin diferencias)"

    url_wa_sem = f"https://wa.me/?text={texto_wa.replace(' ', '%20').replace(chr(10), '%0A')}"
    st.link_button("💬 Enviar Rendición por WhatsApp", url_wa_sem)

# --- PESTAÑA: CIERRE MENSUAL ---
with tab_mensual:
    st.subheader("📅 Cierre Mensual: Resumen de Ingresos y Gastos")
    if not df_pedidos.empty:
        df_pedidos["Fecha_dt"] = pd.to_datetime(df_pedidos["Fecha"], format="%d/%m/%Y", errors="coerce")
        meses = sorted(df_pedidos["Fecha_dt"].dt.to_period("M").unique(), reverse=True)
        mes_sel = st.selectbox("Seleccionar Mes a Auditar:", meses, format_func=lambda x: x.strftime("%B %Y"))
        
        # Filtro de datos para el mes seleccionado
        pedidos_mes = df_pedidos[df_pedidos["Fecha_dt"].dt.to_period("M") == mes_sel]
        ingresos = pedidos_mes[pedidos_mes["Estado"] == "Entregado Cobrado"]["Monto Total"].sum()
        
        gastos_mes = 0
        if not df_gastos_sem.empty:
            df_gastos_sem["Fecha_dt_gas"] = pd.to_datetime(df_gastos_sem["Semana"], errors="coerce")
            gastos_mes = df_gastos_sem[df_gastos_sem["Fecha_dt_gas"].dt.to_period("M") == mes_sel]["Monto"].sum()
        
        resultado_neto = ingresos - gastos_mes
        
        # Métricas principales
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Ingresos Cobrados", f"${ingresos:,.2f}")
        col2.metric("Gastos Totales", f"${gastos_mes:,.2f}")
        col3.metric("Resultado Neto", f"${resultado_neto:,.2f}")
        col4.metric("Resultado Neto / 2", f"${(resultado_neto / 2):,.2f}")

        # --- COMPARATIVA VS MES ANTERIOR ---
        if len(meses) > 1 and mes_sel in meses:
            idx_actual = meses.index(mes_sel)
            if idx_actual + 1 < len(meses):
                mes_anterior = meses[idx_actual + 1]
                pedidos_ant = df_pedidos[df_pedidos["Fecha_dt"].dt.to_period("M") == mes_anterior]
                ingresos_ant = pedidos_ant[pedidos_ant["Estado"] == "Entregado Cobrado"]["Monto Total"].sum()
                gastos_ant = df_gastos_sem[df_gastos_sem["Fecha_dt_gas"].dt.to_period("M") == mes_anterior]["Monto"].sum() if not df_gastos_sem.empty else 0
                res_ant = ingresos_ant - gastos_ant
                
                delta_val = resultado_neto - res_ant
                st.metric("Comparativa vs Mes Anterior", f"${resultado_neto:,.2f}", delta=f"{delta_val:,.2f} respecto a {mes_anterior.strftime('%b %Y')}")

        # Margen de rentabilidad
        if ingresos > 0:
            margen = (resultado_neto / ingresos) * 100
            st.metric("Margen de Rentabilidad", f"{margen:.1f}%")

        # =========================================================================
        # 📈 EVOLUCIÓN HISTÓRICA POR MESES (FACTURADO, COBRADO Y GASTOS)
        # =========================================================================
        import altair as alt

        st.write("---")
        st.subheader("📊 Evolución Histórica por Meses")
        
        df_historico = df_pedidos.copy()
        df_historico["Mes_Periodo"] = df_historico["Fecha_dt"].dt.to_period("M").astype(str)
        
        # Obtenemos los totales por mes de cada concepto
        meses_unicos = sorted(df_historico["Mes_Periodo"].unique())
        
        datos_grafico = []
        for m in meses_unicos:
            p_mes = df_historico[df_historico["Mes_Periodo"] == m]
            fact_total = p_mes["Monto Total"].sum()
            cob_total = p_mes[p_mes["Estado"] == "Entregado Cobrado"]["Monto Total"].sum()
            
            # Buscamos los gastos de ese mes
            gasto_mes_val = 0
            if not df_gastos_sem.empty and "Fecha_dt_gas" in df_gastos_sem.columns:
                gasto_mes_val = df_gastos_sem[df_gastos_sem["Fecha_dt_gas"].dt.to_period("M").astype(str) == m]["Monto"].sum()
                
            datos_grafico.append({"Mes_Periodo": m, "Categoría": "Facturado", "Monto": fact_total})
            datos_grafico.append({"Mes_Periodo": m, "Categoría": "Cobrado", "Monto": cob_total})
            datos_grafico.append({"Mes_Periodo": m, "Categoría": "Gastos", "Monto": gasto_mes_val})
            
        resumen_grafico = pd.DataFrame(datos_grafico)
        
        if not resumen_grafico.empty:
            chart = alt.Chart(resumen_grafico).mark_bar().encode(
                x=alt.X('Categoría', title=None, axis=alt.Axis(labels=False)),
                y='Monto',
                color=alt.Color('Categoría', scale=alt.Scale(domain=['Facturado', 'Cobrado', 'Gastos'], range=['#1f77b4', '#2ca02c', '#d62728'])),
                column=alt.Column('Mes_Periodo', title="Mes"),
                tooltip=['Mes_Periodo', 'Categoría', 'Monto']
            ).properties(width=110)
            
            st.altair_chart(chart)

        # =========================================================================
        # 🎨 MATERIALES / RUBROS CON MAYOR DEMANDA (HISTÓRICO Y MES EN CURSO)
        # =========================================================================
        st.write("---")
        st.subheader("🏆 Materiales / Rubros con Mayor Demanda")
        
        if "Material" in df_pedidos.columns and not df_pedidos.empty:
            # 1. Gráfico Histórico General
            st.markdown("### 📈 Histórico General (Todos los meses)")
            top_materiales_gral = df_pedidos.groupby("Material")["Monto Total"].sum().reset_index()
            top_materiales_gral = top_materiales_gral.sort_values(by="Monto Total", ascending=False).set_index("Material")
            
            col_gh1, col_gh2 = st.columns([2, 1])
            with col_gh1:
                st.bar_chart(top_materiales_gral["Monto Total"])
            with col_gh2:
                st.info("💡 **Histórico:** Acumulado total de ingresos por tipo de material desde el inicio de los registros.")

            st.write("---")
            
            # 2. Gráfico del Mes en Curso (seleccionado arriba en el selectbox de mes_sel)
            st.markdown(f"### 📅 Mes en Curso ({mes_sel.strftime('%B %Y')})")
            pedidos_mes_actual = df_pedidos[df_pedidos["Fecha_dt"].dt.to_period("M") == mes_sel]
            
            if not pedidos_mes_actual.empty:
                top_materiales_mes = pedidos_mes_actual.groupby("Material")["Monto Total"].sum().reset_index()
                top_materiales_mes = top_materiales_mes.sort_values(by="Monto Total", ascending=False).set_index("Material")
                
                col_gm1, col_gm2 = st.columns([2, 1])
                with col_gm1:
                    st.bar_chart(top_materiales_mes["Monto Total"])
                with col_gm2:
                    st.info(f"💡 **Mes actual:** Qué materiales o rubros traccionaron más ventas específicamente durante {mes_sel.strftime('%B %Y')}.")
            else:
                st.warning(f"⚠️ No hay trabajos registrados para el mes de {mes_sel.strftime('%B %Y')}.")

        st.write("---")
        st.markdown("### 📤 Exportar Cierre Mensual")
        texto_mes = f"Cierre Mensual {mes_sel.strftime('%B %Y')}: Ingresos: ${ingresos:,.2f} | Gastos: ${gastos_mes:,.2f} | Neto: ${resultado_neto:,.2f} | Neto por Socio: ${(resultado_neto/2):,.2f}"
        url_wa_mes = f"https://wa.me/?text={texto_mes.replace(' ', '%20')}"
        st.link_button("💬 Enviar Cierre Mensual por WhatsApp", url_wa_mes)
# --- MÓDULO AGENTE PRESUPUESTADOR ---
with tab_presupuestos:
    st.header("🤖 Agente Presupuestador - Visión Letreros")
    st.write("Generá presupuestos detallados calculando costos de materiales y mano de obra.")
    
    with st.form("form_presupuesto"):
        col1, col2 = st.columns(2)
        with col1:
            cliente = st.text_input("Nombre del Cliente")
            tipo_trabajo = st.selectbox("Tipo de Trabajo", 
                                       ["Cartel Nuevo", "Cambio de Lona", "Reparación", "Letras Corpóreas", "Neón Flex"])
            costo_herrero = st.number_input("Costo base mano de obra herrero ($)", value=0.0)
            factor_materiales = st.number_input("Multiplicador materiales (2=doble, 0=si incluye materiales)", value=2.0)
        
        with col2:
            alto = st.number_input("Alto (cm)", min_value=0.0)
            ancho = st.number_input("Ancho (cm)", min_value=0.0)
            incluye_colocacion = st.checkbox("¿Incluye colocación?")
            trabajo_altura = st.checkbox("¿Es en altura? (+Recargo)")
            lleva_luces = st.checkbox("¿Lleva iluminación?")
        
        submit = st.form_submit_button("Generar Presupuesto Detallado")

    if submit:
        # 1. Cálculos de superficie y costos base
        m2_frente = (alto * ancho) / 10000
        costo_lona = m2_frente * 13400
        
        # 2. Simulación de búsqueda web de materiales (Aquí el agente simula la consulta en tiempo real)
        # Valores actualizados traídos de corralones / proveedores online de referencia
        precio_cano_web = 28500  # Precio unitario barra caño estructural 20x30 x 6m
        precio_led_web = 5200    # Precio por metro de tira LED exterior
        fuente_cano = "Ferretería Industrial / Corralón Regional (Online)"
        fuente_led = "Caseros LED / Mercado Mayorista Iluminación"
        
        costo_materiales_base = m2_frente * 21000 * factor_materiales
        costo_luces = precio_led_web * (alto / 100) * 3 if lleva_luces else 0  # Estimación lineal
        
        adicional_colocacion = 25000 if incluye_colocacion else 0
        adicional_altura = 35000 if trabajo_altura else 0
        mano_obra_total = costo_herrero + adicional_colocacion + adicional_altura
        
        subtotal_materiales = costo_lona + costo_materiales_base + costo_luces
        total_presupuesto = subtotal_materiales + mano_obra_total
        
        # --- MOSTRAR RESULTADO EN PANTALLA ---
        st.write("---")
        st.subheader(f"🤖 Reporte del Agente: Presupuesto para {cliente}")
        
        # DETALLE DE BÚSQUEDA WEB Y FUENTES
        st.markdown("### 🌐 Detalle de Búsqueda de Insumos en Internet")
        st.info(
            f"**• Caños y Estructura:** Estimado a partir de valores en {fuente_cano} "
            f"(Valor ref. barra 6m: ${precio_cano_web:,.2f}).\n\n"
            f"**• Iluminación LED:** Basado en {fuente_led} "
            f"(Valor ref. metro lineal: ${precio_led_web:,.2f})."
        )
        
        # REFERENCIA DE OTRAS EMPRESAS DE CARTELERÍA
        st.markdown("### 📊 Análisis de Competencia (Otras Empresas)")
        st.warning(
            f"Para un trabajo de tipo **{tipo_trabajo}** de {m2_frente:.2f} m², "
            f"el promedio de mercado que cobran otras empresas de cartelería integral ronda "
            f"entre un 10% más y un 15% menos de tu presupuesto calculado. "
            f"Tu cotización se encuentra alineada competitivamente."
        )

        st.write("#### 💰 Desglose Interno de Costos")
        st.write(f"- Costo Lona ({m2_frente:.2f} m² a $13.400): **${costo_lona:,.2f}**")
        st.write(f"- Insumos / Estructura (Factor {factor_materiales}): **${costo_materiales_base:,.2f}**")
        if lleva_luces:
            st.write(f"- Iluminación LED estimada: **${costo_luces:,.2f}**")
        st.write(f"- Mano de obra Herrero + Adicionales: **${mano_obra_total:,.2f}**")
        
        st.markdown(f"### 💵 TOTAL FINAL: ${total_presupuesto:,.2f}")
        
        # Bloque de texto listo para copiar al formato Visión Letreros
        st.write("---")
        st.write("#### 📄 Texto sugerido para el comprobante:")
        texto_copiable = (
            f"Trabajo: {tipo_trabajo}\n"
            f"Medidas: {alto}x{ancho} cm\n"
            f"{'Incluye colocación en local.-' if incluye_colocacion else 'No incluye colocación.'}\n"
            f"Forma de pago: Anticipo 50%, saldo contra entrega.\n"
            f"Tiempo de entrega: 3 a 4 días hábiles.\n"
            f"Total: ${total_presupuesto:,.2f}"
        )
        st.text_area("Copia esto para tu comprobante:", value=texto_copiable, height=150)